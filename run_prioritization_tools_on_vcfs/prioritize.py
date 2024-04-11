import json
import shutil
import sys
import traceback

import requests
import openpyxl
import os
import pysam
import gzip
import matplotlib.pyplot as plt

def load_gene_info():
    map_id = {}
    map_alt = {}

    # Replace 'your_file.gz' with the path to your gzipped TSV file
    file_path = 'references/gene_info.gz'

    # Open the gzipped file in binary mode ('rb')
    with gzip.open(file_path, 'rb') as gzipped_file:
        # Iterate through the lines of the decompressed file
        for i, line in enumerate(gzipped_file):
            if i == 0:
                continue

            # Decode each line to UTF-8 and strip any leading/trailing whitespace
            line = line.decode('utf-8').strip()

            # Process the line as needed
            line = line.split("\t")

            if line[0] != "9606":
                continue

            if line[2].startswith("LOC") and line[2] == "LOC" + line[1]:
                continue

            map_id[int(line[1])] = {"Symbol": line[2], "Alternatives": line[4]}
            map_alt[line[2]] = {"NCBI": int(line[1]), "Alternatives": line[4]}

    return map_id, map_alt

def is_vcf_GRCh37(vcf):
    # Open the VCF file and check if it is GRCh38
    vcf = pysam.VariantFile(vcf)
    for record in vcf:
        if record.chrom.startswith("chr"):
            return False
        else:
            return True

def liftover_vcf(vcf):
    basename = os.path.basename(vcf).replace(".vcf.gz", "")

    os.system(f"bcftools view --regions chr1,chr2,chr3,chr4,chr5,chr6,chr7,chr8,chr9,chr10,chr11,chr12,chr13,chr14,chr15,chr16,chr17,chr18,chr19,chr20,chr21,chr22,chrX,chrY,chrMT -Oz -I {vcf} > {basename}_reduced.vcf.gz")
    os.system(f"bcftools reheader {basename}_reduced.vcf.gz --fai references/Homo_sapiens_assembly38.fasta.fai > {basename}_reh_reduced.vcf.gz")
    ret = os.system(f"gatk LiftoverVcf -I {basename}_reh_reduced.vcf.gz  -O {basename}_lifted.vcf.gz  --CHAIN references/GRCh38_to_GRCh37.chain  --REJECT rejected_variants.vcf  -R references/human_g1k_v37_decoy.fasta")
    if ret != 0:
        raise ValueError("LiftOver failed")

    os.remove(f"{basename}_reduced.vcf.gz")
    os.remove(f"{basename}_reh_reduced.vcf.gz")
    shutil.move(f"{basename}_lifted.vcf.gz", vcf)

def search_desease_gene(gene_desease, genes):
    gene_desease = gene_desease.split("/")[0]

    if gene_desease in genes:
        return genes[gene_desease]
    else:
        # We check alternatives
        if not gene_desease in gene_symbol_to_info:
            return 30000

        gene_alt = gene_symbol_to_info[gene_desease]

        for gene in gene_alt["Alternatives"].split("|"):
            if gene in genes:
                return genes[gene]

        return 30000

def run_amelie(vcf, hpos, gene_desease):
    url = 'https://amelie.stanford.edu/api/vcf_api/'

    basename = os.path.basename(vcf).replace(".vcf.gz","")
    basename += "-amelie.json"

    if not os.path.exists("prio_results/" + basename):

        print("Run Amelie on: " + vcf)

        response = requests.post(
            url,
            verify=False,
            files={'vcfFile': open(vcf, 'rb')},
            data={'dominantAlfqCutoff': 0.1,
                  'alfqCutoff': 3.0,  # min. 0.1; max. 3.0 (percent)
                  'filterByCount': False,
                  'hmctCutoff': 1,
                  'alctCutoff': 3,
                  'patientName': 'Test',
                  'patientSex': 'None',  # or 'FEMALE', or None
                  'phenotypes': hpos})

        with open("prio_results/" + basename, 'w') as json_file:
            json.dump(response.json(), json_file)

    # Calculate the rank of the gene
    genes = {}

    # Open the result json file
    with open("prio_results/" + basename) as json_file:
        data = json.load(json_file)
        for i, gene in enumerate(data):
            genes[gene[0]] = i

    return search_desease_gene(gene_desease, genes)


def run_xrare(vcf, hpos, gene_desease):
    # Generate r_script to process vcf

    basename = os.path.basename(vcf)
    if not os.path.exists("prio_results/" + basename.replace(".vcf.gz", "-xrare.csv")):
        if hpos == "":
            print("Run Xrare: No hpos provided " + vcf,file=sys.stderr)
            return

        # Get the current working directory
        current_directory = os.getcwd()

        # Check if the given path is an absolute path
        if os.path.isabs(vcf):
            # Remove the current working directory from the given path
            vcf_relative_path = os.path.relpath(vcf, current_directory)
        else:
            vcf_relative_path = vcf

        script = f"library(xrare)\n\
    \n\
    vcffile <- \"/data/{vcf_relative_path}\"\n\
    dt <- xrare(vcffile=vcffile, hpoid=\"{hpos}\")\n\
    \
    dt = setorder(dt, -xrare_score)\n\
    dt = dt[, .(CHROM,POS,REF,ALT,xrare_score,symbol,pathoACMG,tagsACMG)]\n\
    \n\
    csvfile <- sub(\"\\\\.vcf\\\\.gz$\", \"_xrare.csv\", vcffile)\n\
    write.csv(dt, csvfile, row.names = FALSE)\n\
    "
        f = open("prio_tools/xrare/r_run_analysis", "w")
        f.write(script)
        f.close()

        # Run command to start the docker
        os.system("cd prio_tools/xrare && ./run_xrare")

        basename = os.path.basename(vcf)
        shutil.move(vcf.replace(".vcf.gz", "_xrare.csv"), "prio_results/" + basename.replace(".vcf.gz", "-xrare.csv"))

    # Calculate the rank of the gene
    genes = {}

    j = -1

    # Open the result tsv file
    with open("prio_results/" + basename.replace(".vcf.gz", "-xrare.csv")) as csv_file:
        for i, line in enumerate(csv_file.readlines()):
            if i == 0:
                continue
            line = line.split(",")
            if not line[5] in genes:
                j += 1
            genes[line[5][1:-1]] = j

    return search_desease_gene(gene_desease, genes)

def run_lirical(vcf,hpos,gene_desease):
    basename = os.path.basename(vcf)
    if not os.path.exists("prio_results/" + basename.replace(".vcf.gz", "-lirical.tsv")):

        # Generate pheno package
        script = f"{{\"subject\": {{\n\
            \"id\": \"Case\"\n\
        }},\n\
        \"phenotypicFeatures\": [{{\n"

        for hpo in hpos.split(";"):
            if hpo == "":
                continue
            script += f"        \"type\": {{\n\
                \"id\": \"{hpo}\"\n\
            }},\n"

        script += f"    }}]\n"
        script += f"    \"htsFiles\":\n\
        [{{\n\
            \"uri\": \"../../{vcf}\",\n\
            \"description\": \"test\",\n\
            \"htsFormat\": \"VCF\",\n\
            \"genomeAssembly\": \"GRCh19\",\n\
            \"individualToSampleIdentifiers\": {{\n\
                \"patient1\": \"NA12345\"\n\
            }}\n\
        }}],\n\
        \"metaData\": {{\n\
            \"createdBy\": \"Peter R.\",\n\
            \"resources\": [{{\n\
                \"id\": \"hp\",\n\
                \"name\": \"human phenotype ontology\",\n\
                \"namespacePrefix\": \"HP\",\n\
                \"url\": \"http://purl.obolibrary.org/obo/hp.owl\",\n\
                \"version\": \"2018-03-08\",\n\
                \"iriPrefix\": \"http://purl.obolibrary.org/obo/HP_\"\n\
            }}]\n\
        }}\n\
    }}"

        f = open("prio_tools/lirical/lirical.json", "w")
        f.write(script)
        f.close()

        # Run command to start the docker
        os.system("cd prio_tools/lirical && ./run_lirical")

        shutil.move("prio_tools/lirical/lirical.tsv", "prio_results/" + basename.replace(".vcf.gz", "-lirical.tsv"))

    # Calculate the rank of the gene
    genes = {}

    j = -1

    # Open the result tsv file
    with open("prio_results/" + basename.replace(".vcf.gz", "-lirical.tsv")) as tsv_file:
        for i, line in enumerate(tsv_file.readlines()):
            if i < 5:
                continue
            line = line.split("\t")

            gene_symbol = gene_id_to_info[int(line[6].split(":")[1])]["Symbol"]  # Replace with your NCBI Gene ID

            if not gene_symbol in genes:
                j += 1
            genes[gene_symbol] = j

    return search_desease_gene(gene_desease, genes)



def run_exomiser(vcf,hpos,gene_desease):
    basename = os.path.basename(vcf)
    if not os.path.exists("prio_results/" + basename.replace(".vcf.gz", "-exomiser.json")):

        os.system(f"bcftools query --list-samples  {vcf} > out")
        f = open("out", "r")
        case = f.readline().strip()
        f.close()
        script = f"\n\
id: analysis\n\
subject:\n\
  id: {case}\n\
  sex: MALE\n\
phenotypicFeatures:\n"

        for hpo in hpos.split(";"):
            if hpo == "":
                continue
            script += f"  - type:\n\
      id: {hpo}\n"

        script += f"htsFiles:\n\
  - uri: ../../../{vcf}\n\
    htsFormat: VCF\n\
    genomeAssembly: hg19\n\
metaData:\n\
  created: '2019-11-12T13:47:51.948Z'\n\
  createdBy: julesj\n\
  resources:\n\
    - id: hp\n\
      name: human phenotype ontology\n\
      url: http://purl.obolibrary.org/obo/hp.owl\n\
      version: hp/releases/2019-11-08\n\
      namespacePrefix: HP\n\
      iriPrefix: 'http://purl.obolibrary.org/obo/HP_'\n\
      phenopacketSchemaVersion: 1.0"

        f = open("prio_tools/exomizer/exomiser-cli-13.2.1/exomizer.yml", "w")
        f.write(script)
        f.close()

        # Run command to start exomizer
        try:
            # Use shutil.rmtree to remove the folder and its contents
            shutil.rmtree("prio_tools/exomizer/exomiser-cli-13.2.1/results/")
        except Exception as e:
            pass
        os.system("cd prio_tools/exomizer/exomiser-cli-13.2.1 && java -jar exomiser-cli-13.2.1.jar --sample exomizer.yml --output-directory results")

        shutil.move("prio_tools/exomizer/exomiser-cli-13.2.1/results/" + basename.replace(".vcf.gz", "-exomiser.json"), "prio_results/" + basename.replace(".vcf.gz", "-exomiser.json"))

    # Calculate the rank of the gene
    genes = {}
    #vcf = pysam.VariantFile(vcf)
    #for record in vcf:
    #    genes.append(record.info["GENE"])

    # Open the result json file
    with open("prio_results/" + basename.replace(".vcf.gz", "-exomiser.json")) as json_file:
        data = json.load(json_file)
        for i, gene in enumerate(data):
            genes[gene['geneSymbol']] = i

    return search_desease_gene(gene_desease, genes)

def process_case(xlsx, hpos, gene, frg):
    xlsx_no_extension = os.path.splitext(os.path.basename(xlsx))[0]

    if not os.path.exists("VCFs/" + xlsx_no_extension + ".vcf.gz.tbi"):
        os.system("cd VCFs && tabix -f " + xlsx_no_extension + ".vcf.gz")

    if not os.path.exists("VCFs/" + xlsx_no_extension + "_filtered_sort.vcf.gz"):
        if os.path.exists(xlsx):
            wb_obj = openpyxl.load_workbook(xlsx)
            sheet_obj = wb_obj.active

            chrom_col = 1
            pos_col = 2
            ref_col = 3
            alt_col = 4

            case_vcf = pysam.VariantFile("VCFs/" + xlsx_no_extension + ".vcf.gz")
            header_out = case_vcf.header.copy()
            header_out.add_meta('INFO', items=[('ID', "GENE"), ('Number', '1'), ('Type', 'String'),
                                               ('Description', 'Gene in which the variant is located')])

            case_vcf_out = pysam.VariantFile("VCFs/" + xlsx_no_extension + "_tmp.vcf.gz", "w",
                                             header=header_out)

            for variant in case_vcf.fetch():
                case_vcf_out.write(variant)

            case_vcf_out.close()
            os.system("cd VCFs && tabix -f " + xlsx_no_extension + "_tmp.vcf.gz")

            case_vcf = pysam.VariantFile("VCFs/" + xlsx_no_extension + "_tmp.vcf.gz")
            case_vcf_out = pysam.VariantFile("VCFs/" + xlsx_no_extension + "_filtered.vcf.gz", "w",
                                             header=header_out)

            for row in sheet_obj.iter_rows(min_row=int(2)):
                cell_obj = sheet_obj.cell(row=row[0].row, column=chrom_col)
                chromosome = cell_obj.value

                # Search the VCF in all the seq lab
                cell_obj = sheet_obj.cell(row=row[0].row, column=pos_col)
                position = int(cell_obj.value) - 1

                cell_obj = sheet_obj.cell(row=row[0].row, column=ref_col)
                ref = cell_obj.value

                cell_obj = sheet_obj.cell(row=row[0].row, column=alt_col)
                alt = cell_obj.value

                cell_obj = sheet_obj.cell(row=row[0].row, column=23)
                gene = cell_obj.value

                for variant in case_vcf.fetch(contig=chromosome[3:], start=position, stop=position + 1):
                    if variant.chrom == chromosome[3:] and variant.pos == position + 1 and variant.ref == ref:
                        for valt in variant.alts:
                            if alt == valt:
                                # Found
                                variant.info['GENE'] = gene
                                case_vcf_out.write(variant)

            case_vcf_out.close()
            os.system("cd VCFs && bcftools sort -Oz " + xlsx_no_extension + "_filtered.vcf.gz > " + xlsx_no_extension + "_filtered_sort.vcf.gz")
            os.remove("VCFs/" + xlsx_no_extension + "_filtered.vcf.gz")
            os.system("cd VCFs && tabix -f " + xlsx_no_extension + "_filtered_sort.vcf.gz")

            os.remove("VCFs/" + xlsx_no_extension + "_tmp.vcf.gz")

        else:
            os.system("bcftools filter -Oz -o VCFs/" + xlsx_no_extension + "_filtered_sort.vcf.gz -e 'EXAC_BEST_AF>0.01 || UK10K_AF>0.01 || AFR_AF>0.01 || AMR_AF>0.01 || EAS_AF>0.01 || EAS_AF>0.01 || EUR_AF>0.01 || SAS_AF>0.01 || 1KG_BEST_AF>0.01' " + "VCFs/" + xlsx_no_extension + ".vcf.gz")



    # Once the vcf is create we construct the prioritixation for the 4 tools

    ra = run_amelie("VCFs/" + xlsx_no_extension + "_filtered_sort.vcf.gz", hpos, gene)+1
    rx = run_xrare("VCFs/" + xlsx_no_extension + "_filtered_sort.vcf.gz", hpos, gene)+1
    re = run_exomiser("VCFs/" + xlsx_no_extension + "_filtered_sort.vcf.gz", hpos, gene)+1
    rl = run_lirical("VCFs/" + xlsx_no_extension + "_filtered_sort.vcf.gz", hpos, gene)+1

    frg.write(xlsx_no_extension + "\t" + gene + "\t" + str(ra) + "\t" + str(rx) + "\t" + str(re) + "\t" + str(rl) + "\n")

# Set the exomizer database path

working_dir = os.getcwd()
working_dir += "/prio_tools/exomizer/exomiser-cli-13.2.1/data"

# Define the file path
file_path = "prio_tools/exomizer/exomiser-cli-13.2.1/application.properties"  # Replace with your actual file path

# Create a list to store the lines
new_lines = []

print("Loading gene info...")
gene_id_to_info, gene_symbol_to_info = load_gene_info()
print("Gene info loaded")

# Open the file in read mode and read all the lines into a list
with open(file_path, 'r') as file:
    lines = file.readlines()

# Iterate through the lines, and replace line 26 with the correct path
for i, line in enumerate(lines, start=1):
    if i == 26:
        new_lines.append("exomiser.data-directory=" + working_dir + "\n")
    else:
        new_lines.append(line)

# Open the file in write mode and write the modified lines back to the file
with open(file_path, 'w') as file:
    file.writelines(new_lines)

# read all samples

if not os.path.exists("prio_results"):
    os.mkdir("prio_results")

f = open("all_cases.csv", "r")

#Check vcf present
for line in f.readlines():
    line = line.split(",")
    if not os.path.exists("VCFs/" + line[1] + ".vcf.gz"):
        print("Missing " + line[1])

f.close()

frg = open("rank_gene.tsv","w")
f = open("all_cases.csv", "r")

for line in f.readlines():
    line = line.split(",")

    if not os.path.exists("VCFs/" + line[1] + ".vcf.gz"):
        print("Missing " + line[1])
        continue

    print("processing " + line[1])

    try:
        if is_vcf_GRCh37("VCFs/" + line[1] + ".vcf.gz") == False:
            print("Lifting over " + line[1] + ".vcf.gz")
            liftover_vcf("VCFs/" + line[1] + ".vcf.gz")

        process_case("all_cases/" + line[1] + ".xlsx", line[2].strip(), line[3].strip(),frg)
        shutil.copy("VCFs/" + line[1] + ".vcf.gz", "vcfs_processed/" + line[1] + ".vcf.gz")
        shutil.copy("VCFs/" + line[1] + ".vcf.gz.tbi", "vcfs_processed/" + line[1] + ".vcf.gz.tbi")
    except Exception as e:
        trace = traceback.format_exc()
        print(trace)
        pass

frg.close()

# plot the results with maplotlib

x = []
y_ra = []
y_rx = []
y_re = []
y_rl = []

tot = 0
for i in range(1, 100):
    y_ra.append(0)
    y_rx.append(0)
    y_re.append(0)
    y_rl.append(0)
    x.append(i)

# Open the file and read its lines into a list
with open("rank_gene.tsv", 'r') as file:
    lines = file.readlines()
tot = len(lines)

# Get the line count by finding the length of the list
line_count = len(lines)

f = open("rank_gene.tsv","r")
for line in f.readlines():
    line = line.split("\t")
    for i in range(int(line[2]),100):
        y_ra[i-1] += 1.0 / tot

    for i in range(int(line[3]), 100):
        y_rx[i-1] += 1.0 / tot

    for i in range(int(line[4]), 100):
        y_re[i-1] += 1.0 / tot

    for i in range(int(line[5]), 100):
        y_rl[i-1] += 1.0 / tot

plt.plot(x, y_ra, linestyle='-', color='b', label='Amelie')
plt.plot(x, y_rx, linestyle='-', color='r', label='Xrare')
plt.plot(x, y_re, linestyle='-', color='g', label='Exomiser')
plt.plot(x, y_rl, linestyle='-', color='k', label='Lirical')
plt.xlabel('rank')
plt.ylabel('Percentage')
plt.title('Ranking of genes')
plt.legend()
plt.savefig('rank_gene.png',dpi=300)
plt.close()